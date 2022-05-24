import yfinance as yf
import pandas as pd
from stockFundamentals import information
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

def buildFile(ticker, company, currentPrice, companyFloat):
    workbook = Workbook()
    sheet = workbook.active

    white_text_large = Font(color='ffffff', size=24)
    white_text = Font(color='ffffff')

    number_of_shares = int(companyFloat)
    current_price = float(currentPrice)

    ''' Add model title and row/column labels '''
    sheet['A1'] = company
    sheet['A1'].font = white_text_large
    sheet['A3'] = 'Share Price Per Year'
    sheet['A3'].font = white_text
    sheet['A4'] = 'Market Cap'
    sheet['A6'] = 'Dividend Discount Model'
    sheet['A6'].font = white_text
    sheet['A8'] = 'Year'
    sheet['A8'].font = white_text
    sheet['A9'] = 'Dividends Per Share'
    sheet['A10'] = 'EPS'
    sheet['A11'] = 'Dividend Percentage Change'
    sheet['A12'] = 'Cost of Equity'
    sheet['A13'] = 'Intrinsic Value'
    sheet['A15'] = 'Dividend Growth Rate'
    sheet['A16'] = 'Number Of Shares'
    sheet['B16'] = number_of_shares

    print('Row Labels Added')
    print('============================================================')

    ''' Add year labels and set the font to white '''
    x = 98
    for i in range(2013, 2025):
        column_name = str(chr(x))
        sheet[column_name+'8'] = i
        sheet[column_name+'8'].font = white_text
        sheet[column_name+'3'].font = white_text
        x+=1
        
    print('Years Added')
    print('============================================================')

    ''' Calculate the dividends per year '''
    share = yf.Ticker(ticker)
    df = share.dividends
    df = df.reset_index()
    df['year'] = df['Date'].dt.year
    df['sumDividends'] = df.groupby(['year'])['Dividends'].transform(sum)
    df = df.drop_duplicates(subset=['year'])

    ''' Add the dividends per year to the excel file on row 9 '''
    x = 98
    for index, row in df.iterrows():
        if(int(row['year']) >= 2013):
            column_name = str(chr(x))
            column_name = column_name+'9'
            sheet[column_name] = row['sumDividends']
            x+=1

    print('Dividend History Added')
    print('============================================================')


    ''' Add the end of year prices '''
    df = share.history(period='10y', interval='1d')
    df = df.reset_index()
    df['year'] = df['Date'].dt.year
    res = df.drop_duplicates(subset=['year'])
    res = res[['year', 'Close']]

    ''' Add end of year prices on row 3 and calculate the corresponding market cap on row 4 below the prices '''
    x = 98
    for index, row in res.iterrows():
        if(int(row['year']) >= 2014):
            column_name = str(chr(x))
            sheet[column_name+'3'] = row['Close']
            sheet[column_name+'4'] = row['Close']*number_of_shares
            x+=1

    print('End of Year Prices Added')
    print('============================================================')

    ''' Use the stockFundamentals module to create an information object related to the specified stock '''
    data = information(ticker, company)
    ''' Get the eps for the specified stock '''
    eps = data.earnings_per_share()

    ''' Add the eps for each year to the excel file '''
    x = 98
    if('2013' in eps):
        for i in range(2013, 2022):
            column_name = str(chr(x))
            column_name = column_name+'10'
            sheet[column_name] = float(eps[str(i)])
            x+=1

    print('EPS Added')
    print('============================================================')

    ''' Add styling to rows 1, 3, 6, 8 and 4 '''
    x = 97
    for i in range(2013, 2026):
        column_name = str(chr(x))
        sheet[column_name+'1'].fill = PatternFill('solid', start_color='00257d')
        sheet[column_name+'3'].fill = PatternFill('solid', start_color='00257d')
        sheet[column_name+'6'].fill = PatternFill('solid', start_color='00257d')
        sheet[column_name+'8'].fill = PatternFill('solid', start_color='00257d')
        sheet[column_name+'4'].fill = PatternFill('solid', start_color='7b95d1')
        x+=1

    ''' Add specific styling to the below three cells within the excel file '''
    sheet['B11'].fill = PatternFill('solid', start_color='5c5c5c')
    sheet['B12'].fill = PatternFill('solid', start_color='5c5c5c')
    sheet['B13'].fill = PatternFill('solid', start_color='5c5c5c')

    print('Styling Added')
    print('============================================================')

    ''' Calculate the dividend percentage change from 2013 -> 2022 and add it to the file on row 11 '''
    x = 99
    for i in range(2013, 2022):
        column_name = str(chr(x))
        x+=1
        sheet[column_name+'11'] = (sheet[str(chr(x-1)).capitalize()+'9'].value-sheet[str(chr(x-2)).capitalize()+'9'].value)/sheet[str(chr(x-2)).capitalize()+'9'].value

    print('Dividend Percentage Change Added')
    print('============================================================')

    ''' Add formulas to specific cells including averages, forcasts and sums '''
    sheet['B15'] = '=AVERAGE(C11:K11)'
    sheet['K3'] = current_price
    sheet['B15'] = '=AVERAGE(C11:K11)'
    sheet['L3'] = '=_xlfn.FORECAST.LINEAR(L8,B3:K3,B8:K8)'
    sheet['M3'] = '=_xlfn.FORECAST.LINEAR(M8,C3:L3,C8:L8)'
    sheet['K10'] = '=_xlfn.FORECAST.LINEAR(K8,C10:J10,C8:J8)'
    sheet['L10'] = '=_xlfn.FORECAST.LINEAR(L8,D10:K10,D8:K8)'
    sheet['M10'] = '=_xlfn.FORECAST.LINEAR(M8,E10:L10,E8:L8)'
    sheet['L9'] = '=_xlfn.FORECAST.LINEAR(L8,B9:K9,B8:K8)'
    sheet['M9'] = '=_xlfn.FORECAST.LINEAR(M8,C9:L9,C8:L8)'
    sheet['C12'] = '=((D9/C4)+$B$15)'
    sheet['C13'] = '=_xlfn.SUM((D9/(1+C12))+(D3/(1+C12)))'
    sheet['L11'] = '=(L9-K9)/K9'
    sheet['K4'] = '=_xlfn.SUM(K3*$B$16)'

    print('Formulas Added')
    print('============================================================')

    workbook.save(filename='financial_model.xlsx')
    print('*********************** File Created ***********************')
